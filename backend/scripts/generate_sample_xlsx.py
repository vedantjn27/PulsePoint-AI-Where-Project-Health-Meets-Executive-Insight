"""
Generate 6 sample XLSX files for manual project uploads:
  - 2 RED   (critically at-risk projects)
  - 2 AMBER (projects needing attention)
  - 2 GREEN (healthy, on-track projects)

Each file contains 5 sheets: Summary, Tasks, Milestones, Risks, ScopeChanges
All RAG-scoring fields are included:
  schedule slippage, budget burn, milestone health, blockers, stakeholder sentiment

Run from the backend/ directory:
  .venv\\Scripts\\python.exe scripts/generate_sample_xlsx.py
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).resolve().parents[1] / "sample_data" / "manual_upload_templates"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TODAY = date.today()

# ── Colour helpers ──────────────────────────────────────────────────────────
RED_FILL   = PatternFill("solid", fgColor="C0392B")
AMBER_FILL = PatternFill("solid", fgColor="E67E22")
GREEN_FILL = PatternFill("solid", fgColor="27AE60")
HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
ALT_FILL    = PatternFill("solid", fgColor="ECF0F1")

WHITE_FONT  = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT   = Font(bold=True)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)

THIN = Side(style="thin", color="BDC3C7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _style_header_row(ws, row_idx: int, num_cols: int, fill: PatternFill = HEADER_FILL) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _style_data_row(ws, row_idx: int, num_cols: int, alternate: bool = False) -> None:
    fill = ALT_FILL if alternate else PatternFill("solid", fgColor="FFFFFF")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.alignment = LEFT
        cell.border = BORDER


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 50)


def _write_summary(ws, data: dict) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    headers = ["Field", "Value"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER

    for ri, (k, v) in enumerate(data.items(), 2):
        ws.cell(row=ri, column=1, value=k).border = BORDER
        ws.cell(row=ri, column=1).font = BOLD_FONT
        ws.cell(row=ri, column=2, value=v).border = BORDER
        _style_data_row(ws, ri, 2, alternate=(ri % 2 == 0))


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header_row(ws, 1, len(headers))

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
        _style_data_row(ws, ri, len(headers), alternate=(ri % 2 == 0))
    _auto_width(ws)


# ═══════════════════════════════════════════════════════════════════════════
#  PROJECT DEFINITIONS
# ═══════════════════════════════════════════════════════════════════════════

PROJECTS = [
    # ── GREEN 1 ─────────────────────────────────────────────────────────────
    {
        "filename": "GREEN_1_Digital_Transformation.xlsx",
        "rag": "Green",
        "summary": {
            "Project Name":          "Digital Transformation Initiative",
            "Project ID":            "proj_green_001",
            "Client Name":           "Apex Retail Group",
            "Project Manager":       "Sarah Mitchell",
            "Project Start Date":    TODAY - timedelta(days=42),
            "Project End Date":      TODAY + timedelta(days=88),
            "Budget Total":          500000,
            "Budget Spent":          165000,
            "Actual % Complete":     38,
            "Schedule Slippage (d)": 0,
            "Stakeholder Sentiment": "Positive – sponsor engaged, steering committee satisfied",
            "Overall Status Notes":  "Project is on track. All milestones met on time. Team morale is high, stakeholders are responsive, and blockers are resolved quickly. Budget burn is healthy at 33% spend for 38% completion.",
        },
        "tasks": [
            ["Discovery & Requirements", TODAY - timedelta(days=42), TODAY - timedelta(days=22), 100, "completed",   "Yes"],
            ["Architecture Design",      TODAY - timedelta(days=21), TODAY - timedelta(days=8),  100, "completed",   "Yes"],
            ["Dev Sprint 1",             TODAY - timedelta(days=7),  TODAY + timedelta(days=14), 65,  "in progress", "Yes"],
            ["Dev Sprint 2",             TODAY + timedelta(days=15), TODAY + timedelta(days=42), 0,   "not started", "No"],
            ["UAT",                      TODAY + timedelta(days=43), TODAY + timedelta(days=63), 0,   "not started", "Yes"],
            ["Training & Rollout",       TODAY + timedelta(days=64), TODAY + timedelta(days=88), 0,   "not started", "No"],
        ],
        "milestones": [
            ["Requirements Sign-off",   TODAY - timedelta(days=22), "completed",  "Yes"],
            ["Architecture Approved",   TODAY - timedelta(days=8),  "completed",  "Yes"],
            ["Sprint 1 Demo",           TODAY + timedelta(days=14), "on track",   "Yes"],
            ["UAT Entry Gate",          TODAY + timedelta(days=43), "not started","Yes"],
            ["Go-Live",                 TODAY + timedelta(days=88), "not started","Yes"],
        ],
        "risks": [
            ["Minor SME availability pressure during Sprint 2 planning", "Low",    TODAY - timedelta(days=4),  None,        "open",   "Scheduler proactively aligned SMEs"],
            ["Vendor API documentation slightly delayed",                 "Low",    TODAY - timedelta(days=8),  None,        "open",   "Vendor confirmed delivery by next sprint"],
        ],
        "scope_changes": [
            ["Added dashboard export feature per client request", TODAY - timedelta(days=15), "Low – 3 dev-days added, absorbed in sprint buffer"],
        ],
    },

    # ── GREEN 2 ─────────────────────────────────────────────────────────────
    {
        "filename": "GREEN_2_CRM_Upgrade.xlsx",
        "rag": "Green",
        "summary": {
            "Project Name":          "CRM Platform Upgrade",
            "Project ID":            "proj_green_002",
            "Client Name":           "Solaris Financial",
            "Project Manager":       "James Okonkwo",
            "Project Start Date":    TODAY - timedelta(days=30),
            "Project End Date":      TODAY + timedelta(days=60),
            "Budget Total":          280000,
            "Budget Spent":          72000,
            "Actual % Complete":     32,
            "Schedule Slippage (d)": 0,
            "Stakeholder Sentiment": "Positive – weekly updates well received",
            "Overall Status Notes":  "Delivery is ahead of planned schedule by 2 days. Budget is well managed. All critical milestones on track. No blockers. Stakeholders are highly satisfied.",
        },
        "tasks": [
            ["Current State Assessment", TODAY - timedelta(days=30), TODAY - timedelta(days=18), 100, "completed",   "Yes"],
            ["Data Migration Mapping",   TODAY - timedelta(days=17), TODAY - timedelta(days=5),  100, "completed",   "Yes"],
            ["CRM Configuration",        TODAY - timedelta(days=4),  TODAY + timedelta(days=20), 55,  "in progress", "Yes"],
            ["Integration Testing",      TODAY + timedelta(days=21), TODAY + timedelta(days=42), 0,   "not started", "Yes"],
            ["User Acceptance Testing",  TODAY + timedelta(days=43), TODAY + timedelta(days=55), 0,   "not started", "No"],
            ["Go-Live Preparation",      TODAY + timedelta(days=56), TODAY + timedelta(days=60), 0,   "not started", "Yes"],
        ],
        "milestones": [
            ["Assessment Complete",     TODAY - timedelta(days=18), "completed",   "Yes"],
            ["Data Map Approved",       TODAY - timedelta(days=5),  "completed",   "Yes"],
            ["Config Review",           TODAY + timedelta(days=20), "on track",    "Yes"],
            ["Integration Gate",        TODAY + timedelta(days=42), "not started", "Yes"],
            ["Go-Live",                 TODAY + timedelta(days=60), "not started", "Yes"],
        ],
        "risks": [
            ["Potential delay in test data provisioning", "Low", TODAY - timedelta(days=3), None, "open", "Test data request submitted; due in 2 days"],
        ],
        "scope_changes": [],
    },

    # ── AMBER 1 ─────────────────────────────────────────────────────────────
    {
        "filename": "AMBER_1_ERP_Rollout.xlsx",
        "rag": "Amber",
        "summary": {
            "Project Name":          "ERP Global Rollout",
            "Project ID":            "proj_amber_001",
            "Client Name":           "Helio Manufacturing",
            "Project Manager":       "Marcus Chen",
            "Project Start Date":    TODAY - timedelta(days=75),
            "Project End Date":      TODAY + timedelta(days=55),
            "Budget Total":          650000,
            "Budget Spent":          310000,
            "Actual % Complete":     45,
            "Schedule Slippage (d)": 8,
            "Stakeholder Sentiment": "Mixed – finance team concerned about data quality; IT sponsor supportive",
            "Overall Status Notes":  "Project is recoverable but needs management attention. Data quality issues from legacy system are slowing the migration phase. One milestone is delayed. Budget is slightly over-burned at 48% spend for 45% completion. Escalation may be needed if data remediation extends beyond next sprint.",
        },
        "tasks": [
            ["Requirements & Blueprint",  TODAY - timedelta(days=75), TODAY - timedelta(days=48), 100, "completed",   "Yes"],
            ["System Configuration",      TODAY - timedelta(days=47), TODAY - timedelta(days=12), 90,  "completed",   "Yes"],
            ["Data Migration – Extract",  TODAY - timedelta(days=11), TODAY + timedelta(days=8),  60,  "at risk",     "Yes"],
            ["Data Migration – Load",     TODAY + timedelta(days=9),  TODAY + timedelta(days=30), 0,   "not started", "Yes"],
            ["Integration Testing",       TODAY + timedelta(days=31), TODAY + timedelta(days=48), 0,   "not started", "Yes"],
            ["Cutover & Hypercare",       TODAY + timedelta(days=49), TODAY + timedelta(days=55), 0,   "not started", "Yes"],
        ],
        "milestones": [
            ["Blueprint Approved",        TODAY - timedelta(days=48), "completed",  "Yes"],
            ["Config Freeze",             TODAY - timedelta(days=12), "completed",  "Yes"],
            ["Mock Migration 1",          TODAY - timedelta(days=3),  "delayed",    "Yes"],
            ["Mock Migration 2",          TODAY + timedelta(days=15), "at risk",    "Yes"],
            ["Integration Test Entry",    TODAY + timedelta(days=31), "not started","Yes"],
            ["Go-Live Readiness",         TODAY + timedelta(days=55), "not started","Yes"],
        ],
        "risks": [
            ["Legacy data quality issues causing extraction rework",              "High",   TODAY - timedelta(days=18), None, "open",   "Data cleansing team engaged; ETA +12 days"],
            ["Key SME unavailability during data load phase",                     "Medium", TODAY - timedelta(days=9),  None, "open",   "Backup resource being arranged"],
            ["Third-party integration vendor response time is slow",              "Medium", TODAY - timedelta(days=5),  None, "open",   "Escalated to vendor account manager"],
        ],
        "scope_changes": [
            ["Additional legacy chart-of-accounts mapping added",  TODAY - timedelta(days=14), "Moderate – 8 dev-days, partially absorbing buffer"],
            ["New regulatory reporting requirement added by client", TODAY - timedelta(days=6), "Low – handled within existing sprint"],
        ],
    },

    # ── AMBER 2 ─────────────────────────────────────────────────────────────
    {
        "filename": "AMBER_2_Supply_Chain_Optimisation.xlsx",
        "rag": "Amber",
        "summary": {
            "Project Name":          "Supply Chain Optimisation",
            "Project ID":            "proj_amber_002",
            "Client Name":           "NovaTrade Logistics",
            "Project Manager":       "Priya Nair",
            "Project Start Date":    TODAY - timedelta(days=60),
            "Project End Date":      TODAY + timedelta(days=40),
            "Budget Total":          390000,
            "Budget Spent":          195000,
            "Actual % Complete":     48,
            "Schedule Slippage (d)": 5,
            "Stakeholder Sentiment": "Neutral – operations team cautious; exec sponsor supportive",
            "Overall Status Notes":  "Project is under moderate pressure. Budget burn is at 50% spend against 48% completion — slightly over. One near-term milestone is at risk due to integration delays with a third-party logistics API. Team is managing but needs close monitoring.",
        },
        "tasks": [
            ["Process Mapping",          TODAY - timedelta(days=60), TODAY - timedelta(days=38), 100, "completed",   "No"],
            ["Vendor Selection",         TODAY - timedelta(days=37), TODAY - timedelta(days=20), 100, "completed",   "No"],
            ["Platform Configuration",   TODAY - timedelta(days=19), TODAY + timedelta(days=5),  75,  "in progress", "Yes"],
            ["3PL API Integration",      TODAY - timedelta(days=5),  TODAY + timedelta(days=15), 30,  "at risk",     "Yes"],
            ["Pilot & UAT",              TODAY + timedelta(days=16), TODAY + timedelta(days=35), 0,   "not started", "Yes"],
            ["Rollout",                  TODAY + timedelta(days=36), TODAY + timedelta(days=40), 0,   "not started", "No"],
        ],
        "milestones": [
            ["Vendor Contract Signed",   TODAY - timedelta(days=20), "completed",  "No"],
            ["Config Sign-off",          TODAY + timedelta(days=5),  "on track",   "Yes"],
            ["Integration Test Gate",    TODAY + timedelta(days=15), "at risk",    "Yes"],
            ["UAT Entry",               TODAY + timedelta(days=16), "not started", "Yes"],
            ["Go-Live",                  TODAY + timedelta(days=40), "not started", "Yes"],
        ],
        "risks": [
            ["3PL API credentials delayed by logistics partner",          "Medium", TODAY - timedelta(days=12), None, "open",   "Follow-up meeting scheduled with partner"],
            ["Pilot team availability constrained during school holidays", "Low",    TODAY - timedelta(days=4),  None, "open",   "Alternative pilot dates identified"],
        ],
        "scope_changes": [
            ["Real-time tracking dashboard added to scope", TODAY - timedelta(days=10), "Moderate – 5 dev-days added"],
        ],
    },

    # ── RED 1 ────────────────────────────────────────────────────────────────
    {
        "filename": "RED_1_Vendor_Portal_Recovery.xlsx",
        "rag": "Red",
        "summary": {
            "Project Name":          "Vendor Portal Recovery Programme",
            "Project ID":            "proj_red_001",
            "Client Name":           "Contoso Enterprises",
            "Project Manager":       "Alex Morgan",
            "Project Start Date":    TODAY - timedelta(days=90),
            "Project End Date":      TODAY + timedelta(days=20),
            "Budget Total":          220000,
            "Budget Spent":          198000,
            "Actual % Complete":     50,
            "Schedule Slippage (d)": 22,
            "Stakeholder Sentiment": "Negative – client has escalated twice; exec sponsor demanding weekly steering",
            "Overall Status Notes":  "Project is in critical condition. Budget is 90% consumed with only 50% completion. Schedule has slipped 22 days. A critical blocker on third-party API credentials has been open for 25 days. Two milestones are overdue. Formal recovery plan required immediately. Escalation to executive steering committee is recommended.",
        },
        "tasks": [
            ["Discovery & Scoping",      TODAY - timedelta(days=90), TODAY - timedelta(days=60), 100, "completed",  "No"],
            ["API Integration Design",   TODAY - timedelta(days=59), TODAY - timedelta(days=30), 90,  "at risk",    "Yes"],
            ["Backend Development",      TODAY - timedelta(days=29), TODAY + timedelta(days=5),  55,  "blocked",    "Yes"],
            ["Frontend Development",     TODAY - timedelta(days=15), TODAY + timedelta(days=10), 40,  "blocked",    "Yes"],
            ["Security Review",          TODAY + timedelta(days=6),  TODAY + timedelta(days=16), 0,   "not started","Yes"],
            ["Deployment & Cutover",     TODAY + timedelta(days=17), TODAY + timedelta(days=20), 0,   "not started","Yes"],
        ],
        "milestones": [
            ["API Design Approved",      TODAY - timedelta(days=30), "delayed",  "Yes"],
            ["Backend Dev Complete",     TODAY - timedelta(days=5),  "at risk",  "Yes"],
            ["Security Clearance",       TODAY + timedelta(days=16), "at risk",  "Yes"],
            ["Cutover Readiness",        TODAY + timedelta(days=20), "at risk",  "Yes"],
        ],
        "risks": [
            ["Third-party API credentials not issued by vendor",               "Critical", TODAY - timedelta(days=25), None, "open",   "Vendor legal review blocking; legal team escalating"],
            ["Security audit team backlog delaying review start",              "High",     TODAY - timedelta(days=18), None, "open",   "Alternative audit firm being evaluated"],
            ["Key backend developer on sick leave for 2 weeks",               "High",     TODAY - timedelta(days=10), None, "open",   "Contractor being sourced urgently"],
            ["Client stakeholder demanding scope extension mid-recovery",      "Medium",   TODAY - timedelta(days=5),  None, "open",   "Scope freeze discussion required in steering"],
        ],
        "scope_changes": [
            ["Vendor security review added late to project scope",        TODAY - timedelta(days=12), "High – 10 days delay risk on critical path"],
            ["Additional reporting module requested by client post-UAT",  TODAY - timedelta(days=4),  "High – 8 dev-days, no buffer remaining"],
        ],
    },

    # ── RED 2 ────────────────────────────────────────────────────────────────
    {
        "filename": "RED_2_Core_Banking_Migration.xlsx",
        "rag": "Red",
        "summary": {
            "Project Name":          "Core Banking System Migration",
            "Project ID":            "proj_red_002",
            "Client Name":           "Meridian Bank",
            "Project Manager":       "Diana Fernandez",
            "Project Start Date":    TODAY - timedelta(days=120),
            "Project End Date":      TODAY + timedelta(days=10),
            "Budget Total":          900000,
            "Budget Spent":          880000,
            "Actual % Complete":     65,
            "Schedule Slippage (d)": 35,
            "Stakeholder Sentiment": "Negative – board-level concern; CIO considering project suspension",
            "Overall Status Notes":  "Project is critically over budget and severely delayed. Budget is 98% consumed with only 65% completion. Schedule has slipped 35 days from original baseline. Critical regulatory testing milestone is 3 weeks overdue. Multiple blockers including data centre access issues and regulatory sign-off delays. Immediate executive intervention required.",
        },
        "tasks": [
            ["Requirements & Regulatory Mapping",  TODAY - timedelta(days=120), TODAY - timedelta(days=85), 100, "completed",  "Yes"],
            ["Legacy Data Extraction",             TODAY - timedelta(days=84),  TODAY - timedelta(days=50), 100, "completed",  "Yes"],
            ["Core Banking Config",                TODAY - timedelta(days=49),  TODAY - timedelta(days=10), 85,  "at risk",    "Yes"],
            ["Regulatory UAT",                     TODAY - timedelta(days=9),   TODAY + timedelta(days=5),  30,  "blocked",    "Yes"],
            ["Parallel Run",                       TODAY + timedelta(days=6),   TODAY + timedelta(days=10), 0,   "not started","Yes"],
            ["Cutover",                            TODAY + timedelta(days=10),  TODAY + timedelta(days=10), 0,   "not started","Yes"],
        ],
        "milestones": [
            ["Data Extraction Complete",        TODAY - timedelta(days=50), "completed",  "Yes"],
            ["Core Config Freeze",              TODAY - timedelta(days=10), "delayed",    "Yes"],
            ["Regulatory UAT Sign-off",         TODAY - timedelta(days=2),  "at risk",    "Yes"],
            ["Parallel Run Approval",           TODAY + timedelta(days=8),  "at risk",    "Yes"],
            ["Cutover Approved",                TODAY + timedelta(days=10), "at risk",    "Yes"],
        ],
        "risks": [
            ["Regulatory body sign-off delayed – new requirements surfaced",  "Critical", TODAY - timedelta(days=22), None, "open",   "Regulator liaison engaged; response expected +10d"],
            ["Data centre access restrictions preventing parallel run setup", "Critical", TODAY - timedelta(days=15), None, "open",   "Data centre team escalated to CTO"],
            ["Core banking vendor support SLA breached twice",               "High",     TODAY - timedelta(days=30), None, "open",   "Vendor account director on call"],
            ["Staff fatigue and attrition risk after extended overtime",      "High",     TODAY - timedelta(days=8),  None, "open",   "HR reviewing retention measures"],
            ["Incomplete test scripts for Parallel Run phase",               "Medium",   TODAY - timedelta(days=5),  None, "open",   "Test lead brought in to fast-track"],
        ],
        "scope_changes": [
            ["New anti-money laundering module mandated by regulator",   TODAY - timedelta(days=25), "Critical – 3-week delay, $80k additional cost"],
            ["Real-time fraud scoring integration added by client",       TODAY - timedelta(days=18), "High – 2-week delay, $40k additional cost"],
            ["Additional data archival requirements added",               TODAY - timedelta(days=7),  "Medium – 5 dev-days"],
        ],
    },
]

TASK_HEADERS      = ["Task Name", "Start Date", "End Date", "% Complete", "Status", "Critical Path", "Notes"]
MILESTONE_HEADERS = ["Milestone Name", "Due Date", "Status", "Critical Path", "Notes"]
RISK_HEADERS      = ["Risk / Blocker Description", "Severity", "Opened Date", "Resolved Date", "Status", "Mitigation / Notes"]
SCOPE_HEADERS     = ["Scope Change Description", "Change Date", "Impact", "Approved By", "Notes"]


def build_workbook(proj: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ───────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _write_summary(ws_sum, proj["summary"])

    # ── Sheet 2: Tasks ─────────────────────────────────────────────────────
    ws_tasks = wb.create_sheet("Tasks")
    task_rows = []
    for t in proj["tasks"]:
        task_rows.append([t[0], t[1].isoformat(), t[2].isoformat(), t[3], t[4], t[5], ""])
    _write_sheet(ws_tasks, TASK_HEADERS, task_rows)

    # ── Sheet 3: Milestones ────────────────────────────────────────────────
    ws_ms = wb.create_sheet("Milestones")
    ms_rows = []
    for m in proj["milestones"]:
        ms_rows.append([m[0], m[1].isoformat(), m[2], m[3], ""])
    _write_sheet(ws_ms, MILESTONE_HEADERS, ms_rows)

    # ── Sheet 4: Risks ─────────────────────────────────────────────────────
    ws_risks = wb.create_sheet("Risks")
    risk_rows = []
    for r in proj["risks"]:
        resolved = r[3].isoformat() if r[3] else ""
        risk_rows.append([r[0], r[1], r[2].isoformat(), resolved, r[4], r[5]])
    _write_sheet(ws_risks, RISK_HEADERS, risk_rows)

    # ── Sheet 5: ScopeChanges ──────────────────────────────────────────────
    ws_sc = wb.create_sheet("ScopeChanges")
    sc_rows = []
    for s in proj["scope_changes"]:
        sc_rows.append([s[0], s[1].isoformat(), s[2], "", ""])
    _write_sheet(ws_sc, SCOPE_HEADERS, sc_rows)

    return wb


def main() -> None:
    rag_fill = {"Red": RED_FILL, "Amber": AMBER_FILL, "Green": GREEN_FILL}

    for proj in PROJECTS:
        wb = build_workbook(proj)

        # Colour the Summary sheet tab
        fill_colour = rag_fill[proj["rag"]].fgColor.rgb
        wb["Summary"].sheet_properties.tabColor = fill_colour

        out_path = OUTPUT_DIR / proj["filename"]
        wb.save(out_path)
        print(f"  [OK]  {proj['filename']}")

    print(f"\nAll 6 files saved to:\n  {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
